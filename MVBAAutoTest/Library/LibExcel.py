from io import BytesIO
from Params.SystemParams import FileExt
from openpyxl import *
from openpyxl.drawing.image import Image
import os


class LibExcel:
    """Excel幫助類"""
# region Property

# endregion

# region Construct
    def __init__(self, path, name):
        self.Path = path
        self.fileName = name
        self.__workbook: Workbook = None
    # region VirtualFunction
        self._AfterSetData = None
    # endregion
# endregion

# region Protected
    def _Createfile(self, isCreateNew=False):
        if os.path.isfile(self.Path + self.fileName):
            if isCreateNew:
                '覆蓋'
            else:
                self.__workbook = self._Openfile()
        else:
            self.__workbook = Workbook()

    def _Openfile(self, srcPath, srcFileName):
        if(not os.path.isdir(srcPath)):
            "報錯：找不到資料夾"
            return
        if(not os.path.isfile(srcPath+'/'+srcFileName)):
            "報錯：找不到"
            return
        self.__workbook = load_workbook(srcPath+'/'+srcFileName)

    def _WriteDatas(self, sheets=[], isWriteEmpty=False):
        idx = 0
        for table in sheets:
            self._WriteData(idx, table, sheets[table], isWriteEmpty)
            idx += 1

    def _WriteData(self, sheetIdx, sheetName, table=[], isWriteEmpty=False):
        if(sheetName in self.__workbook.sheetnames):
            self.__workbook.active = self.__workbook.get_sheet_by_name(
                sheetName)
        else:
            self.__workbook.create_sheet(sheetName, sheetIdx)
            self.__workbook.active = sheetIdx
        sheet = self.__workbook.active
        rowIdx = 1
        for row in table:
            columnIdx = 1
            for columnName, columnValue in row.__dict__.items():
                if(columnValue != None):
                    if(isinstance(columnValue, str)):
                        if(columnValue != '' or isWriteEmpty):
                            sheet.cell(rowIdx, columnIdx, columnValue)
                    elif (isinstance(columnValue, BytesIO)):
                        sheet.add_image(Image(columnValue), chr(
                            columnIdx+64)+str(rowIdx))
                    if(not self._AfterSetData is None):
                        self._AfterSetData(sheet.row_dimensions[rowIdx], sheet.column_dimensions[chr(
                            columnIdx+64)], columnValue)
                columnIdx += 1
            rowIdx += 1

    def _AutoSize(self):
        sheets = self.__workbook.sheetnames
        for sheetName in sheets:
            sheet = self.__workbook[sheetName]
           # self.__InitSize(sheet)
            for row in sheet.rows:
                for cell in row:
                    self.__SetMaxColumnSize(sheet, cell)
                    self.__SetMaxRowSize(sheet, cell)

    def _Save(self):
        self.__workbook.save()
        self.__workbook.close()

    def _SaveAs(self):
        self.__workbook.save(self.Path + "/" + self.fileName + FileExt._xlsx)
        self.__workbook.close()
# endregion

# region Private
    def __InitSize(self, sheet):
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = 0

    # 設置最大欄位寬
    def __SetMaxColumnSize(self, sheet, cell):
        if cell.value:
            sheet.column_dimensions[cell.column_letter].width = max(
                sheet.column_dimensions[cell.column_letter].width, len(str(cell.value).encode('big5'))*1.5)
    # 設置最大行高

    def __SetMaxRowSize(self,  sheet, cell):
        ""
# endregion
