from openpyxl.styles import Alignment, Font
import time
from openpyxl import *
from openpyxl.drawing.image import Image
import openpyxl
from Params.ElementParams import ElementParam

stepNum = 0
max_width = 0

# new_sheet(匯出的excel檔名) ->建立工作表


def new_sheet(fileName, sheetName=''):
    try:
        wb = load_workbook(fileName+'.xlsx')          # 開啟 xlsx
    except:
        wb = Workbook()  # 若檔案不存在則建立新的workbook
    sheet = wb.create_sheet(sheetName, 0)  # 建立新的工作表

    sheet['A1'] = '時間'
    sheet['B1'] = '步驟 / 執行動作'
    sheet['C1'] = '結果'
    sheet['D1'] = '圖片'

    sheet.column_dimensions['A'].width = 18.0  # 調整列寬
    sheet.column_dimensions['B'].width = 50.0
    sheet.column_dimensions['C'].width = 50.0
    #sheet.column_dimensions['D'].width = 48
    wb.save(fileName+'.xlsx')  # 儲存檔案


# export_table(匯出的excel檔名,step(步驟或執行動作),結果,圖片)         ->匯出到excel
def export_table(fileName, w=0, step='', result='', img_path=''):
    # 開啟 xlsx  './+fileName+'.xlsx'
    wb = load_workbook(fileName+'.xlsx')
    wb.active = w
    sheet = wb.active                     # 啟用工作表

    localtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    rowNum = sheet.max_row  # 總行數
    align = Alignment(horizontal='left', vertical='center',
                      wrapText=True)  # 設定字體靠左,置中,自動換行
    sheet.cell(row=rowNum+1, column=1, value=localtime)  # 寫入時間
    sheet['A'+str(rowNum+1)].alignment = align

    sheet.cell(row=rowNum+1, column=2, value=step)  # 寫入步驟
    sheet['B'+str(rowNum+1)].alignment = align

    sheet.cell(row=rowNum+1, column=3, value=result)  # 寫入結果
    sheet['c'+str(rowNum+1)].alignment = align

    if result == 'Failed' or result == "系統運行錯誤，請看log日誌":  # 若結果 = 'Failed',顯示成紅色
        sheet['C'+str(rowNum+1)].font = Font(color='FF0000')
    elif result == 'Pass':
        sheet['C'+str(rowNum+1)].font = Font(color='0000FF')

    if img_path != '':
        img = Image(img_path)
        cell_height = img.height * 3/4  # 圖片長 -> 儲存格 高
        global max_width
        if cell_height < 135:
            sheet.row_dimensions[rowNum+1].height = cell_height
            cell_width = img.width * 3/20  # 圖片寬 -> 儲存格 寬
            # print(cell_width)
            if cell_width > max_width:
                sheet.column_dimensions['D'].width = cell_width
                max_width = cell_width

        else:
            sheet.row_dimensions[rowNum+1].height = 135  # 調整行高
            n = img.height/180
            img.width, img.height = img.width/n, img.height/n  # 調整圖片大小
            cell_width = img.width*3/20
            # print(cell_width)
            if cell_width > max_width:
                sheet.column_dimensions['D'].width = cell_width
                max_width = cell_width
        sheet.add_image(img, 'D'+str(rowNum+1))  # 插入圖片

    else:
        pass
    wb.save(fileName+'.xlsx')  # 儲存檔案


def TestCase_steps(TestCase_path, w=0):  # 讀取Test case步驟
    global stepNum
    workbook = openpyxl.load_workbook(TestCase_path+'.xlsx')
    workbook.active = w
    sheet = workbook.active
    step_list = []
    # print(sheet.max_row)
    for i in range(1, sheet.max_row):
        step = sheet['B'+str(i + 1)].value
        if step != None:
            step_list.append(step)
            # print(step_list)

    # if stepNum == len(step_list):
    #     stepNum =1          #Step 1 (登入) 只會執行一次
    # stepNum +=1
    # #print(stepNum)
    return step_list
