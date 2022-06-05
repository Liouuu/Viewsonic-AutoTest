from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import time
from UnitFunction import LoginAndActive, ImageSearch, MagicBox, log
import sys
import os
sys.path.append(os.getcwd())


ImageSearch_w = Workbook()
w_s = ImageSearch_w.active
w_s.title = "Png"
w_s1 = ImageSearch_w.create_sheet(index=1, title="Svg")
w_s2 = ImageSearch_w.create_sheet(index=2, title="Gif")
for w in range(3):
    sheet = ImageSearch_w.worksheets[w]
    sheet['A1'] = "項目"
    sheet['B1'] = "步驟"
    sheet['C1'] = "結果"
    sheet.column_dimensions['A'].width = 18.0  # 調整列寬
    sheet.column_dimensions['B'].width = 50.0
    sheet.column_dimensions['C'].width = 50.0
    sheet.column_dimensions['D'].width = 48
ImageSearch_w.save('./Output/Excels/ImageSearch.xlsx')
ImageSearch_wd = load_workbook('./Output/Excels/ImageSearch.xlsx')
ImageSearch_log = log.Logger(
    './Output/Logs/ImageSearch'+datetime.datetime.now().strftime(" %Y%m%d %H_%M_%S") + '.log', level='info')


def ImportimagetypeCase(self, type="svg"):  # Image type:"png" , "svg" , "gif"
    if type == "png":
        sheet = ImageSearch_wd.worksheets[0]
    elif type == "svg":
        sheet = ImageSearch_wd.worksheets[1]
    else:
        sheet = ImageSearch_wd.worksheets[2]
    try:
        sheet['B2'] = "Start ImportimagetypeCase"
        sheet['B3'] = "Normal Login"
        LoginAndActive.NormalLogin(
            self, "mvbaautotest2@gmail.com", "$Password1")
        time.sleep(15)
        sheet['B4'] = "Open Magic Box"
        MagicBox.OpenMagicBox(self)
        sheet['B5'] = "Select Image Search"
        MagicBox.SelectImageSearch(self)
        sheet['B6'] = "Searching..."
        resultSearch = ImageSearch.inputsearch(self, type)
        sheet['B7'] = resultSearch
        result, result2 = ImageSearch.Importimage(self, type)
        if result == "搜尋超時，請稍後再試~":
            sheet['C7'] = result
            return
        elif result == True:
            sheet['C7'] = "搜索到Image參數:" + result2
            sheet['C8'] = "搜索到Image檔案格式:" + result2[-3::]
            sheet['B9'] = "Try import image!"
        else:
            sheet['C8'] = result2
            return

        result3, result4 = ImageSearch.Checktype(self, type)
        if result3 == "Import異常，請稍後再試~":
            sheet['C9'] = result3
            return
        elif result3 == True:
            sheet['B10'] = "Check Imported Image type"
            sheet['C9'] = "成功import"
            sheet['C10'] = "成功import的image為:" + result4
            sheet['C11'] = "Check Image type Pass!"
        else:
            sheet['C10'] = "成功import的Image為:" + result4 + "! 與實際Type不匹配，請稍後再試~"
        time.sleep(5)
        sheet['C11'] = "ImportimagetypeCase Finish"
    except (SystemExit, KeyboardInterrupt):
        raise
    except Exception:
        sheet['C2'] = "系統運行錯誤，請看log日誌"
        ImageSearch_log.logger.error("系統運行錯誤，請重試", exc_info=True)
        ImageSearch_log.logger.error("Finish")
    ImageSearch_wd.save('./Excels/ImageSearch.xlsx')
