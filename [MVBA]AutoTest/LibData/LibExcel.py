from openpyxl.drawing.image import Image
from Params. SystemParams import FileExt
from openpyxl import *
from openpyxl.utils import get_column_letter
import os


class LibExcel:
    __workbook = Workbook

    def __init__(self, path, name):
        self.Path = path
        self.fileName = name

    def Createfile(self, isCreateNew=False):
        if os.path.isfile(self.Path + self.fileName):
            if isCreateNew:
                '覆蓋'
            else:
                self.__workbook = self.Openfile()
        else:
            self.__workbook = Workbook()

    def Openfile(self, srcPath, srcFileName):
        if(not os.path.isdir(srcPath)):
            "報錯：找不到資料夾"
            return
        if(not os.path.isfile(srcPath+'/'+srcFileName)):
            "報錯：找不到"
            return
        self.__workbook = load_workbook(srcPath+'/'+srcFileName)

    def WriteDatas(self, sheets=[], isWriteEmpty=False):
        idx = 0
        for table in sheets:
            self.WriteData(idx, table, sheets[table], isWriteEmpty)
            idx += 1

    def WriteData(self, sheetIdx, sheetName, table=[], isWriteEmpty=False):
        if(sheetName in self.__workbook.sheetnames):
            self.__workbook.active = self.__workbook.get_sheet_by_name(
                sheetName)

        else:
            self.__workbook.create_sheet(sheetName, sheetIdx)
            self.__workbook.active = sheetIdx
        sheet = self.__workbook.active
        rowIdx = 1
        for row in table:
            columnIdx = 1
            for columnName, colunmValue in row.__dict__.items():
                if(isinstance(colunmValue, str)):
                    if(colunmValue != '' or isWriteEmpty):
                        sheet.cell(rowIdx, columnIdx, colunmValue)
                elif (isinstance(colunmValue, Image)):
                    # Ascii取大寫英文字母欄位
                    sheet.add_image(colunmValue, chr(columnIdx+64)+str(rowIdx))
                columnIdx += 1
            rowIdx += 1

    def AutoSize(self):
        sheets = self.__workbook.sheetnames
        for sheetName in sheets:
            widths = []
            heights = []
            sheet = self.__workbook[sheetName]
            for row in sheet.rows:
                for cell in row:
                    self.__GetMaxColumnSize(cell, *widths)
                    self.__GetMaxRowSize(cell, heights)
            self.__SetMaxColumnSize()
            self.__SetMaxRowSize()

    def Save(self):
        self.__workbook.save()
        self.__workbook.close()

    def SaveAs(self):
        self.__workbook.save(self.Path + "/" + self.fileName + FileExt._xlsx)
        self.__workbook.close()
    # 獲取最大欄位寬

    def __GetMaxColumnSize(cell, *widths):
        if cell.value:
            widths[cell.column] = max(widths.get(
                cell.column, 0), len(str(cell.value).encode('utf-8')))
    # 獲取最大行高

    def __GetMaxRowSize(cell):
        ""

    # 設置最大欄位寬

    def __SetMaxColumnSize(sheet):
        for col, value in dims.items():
            sheet.column_dimensions[get_column_letter(col)].width = value
    # 設置最大行高

    def __SetMaxRowSize(sheet):
        for col, value in dims.items(sheet):
            sheet.column_dimensions[get_column_letter(col)].width = value
