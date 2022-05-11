from openpyxl import load_workbook
from openpyxl import Workbook
from unittest import result
import time
import datetime
from UnitFunction import log, LoginAndActive, Tier
import sys
import os
sys.path.append(os.getcwd())
Tier_w = Workbook()
w_s = Tier_w.active
w_s.title = "No Activate"
for w_index in range(4):
    w_s1 = Tier_w.create_sheet(index=w_index+1, title="Tier"+str(w_index+1))
for w in range(5):
    sheet = Tier_w.worksheets[w]
    sheet['A1'] = "項目"
    sheet['B1'] = "步驟"
    sheet['C1'] = "結果"
    sheet.column_dimensions['A'].width = 18.0  # 調整列寬
    sheet.column_dimensions['B'].width = 50.0
    sheet.column_dimensions['C'].width = 50.0
    sheet.column_dimensions['D'].width = 48
Tier_w.save('./Output/Excels/Tier.xlsx')
Tier_wd = load_workbook('./Output/Excels/Tier.xlsx')
Tier_log = log.Logger(
    './Output/Logs/Tier'+datetime.datetime.now().strftime(' %Y%m%d %H_%M_%S') + '.log', level='info')


def NoActivateCase(self):
    #Tier_wd = load_workbook('./Excels/Tier.xlsx')
    sheet = Tier_wd.worksheets[0]
    try:
        sheet['B2'] = "Check WirelessPresentation for No Activate"
        if Tier.TryWirelessPresentation(self)[1] == "Whiteboard activation required" or "啟用提示":
            # print(Tier.TryWirelessPresentation(self)[1])
            sheet['C2'] = "WirelessPresentation for No Activate is pass!!"
        else:
            # print(Tier.TryWirelessPresentation(self)[0])
            sheet['C2'] = "WirelessPresentation for No Activate is fail!!"

        sheet['B3'] = "Check CaptureAll for No Activate"
        if Tier.TryCaptureAll(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C3'] = "CaptureAll for No Activate is pass!!"
        else:
            sheet['C3'] = "CaptureAll for No Activate is fail!!"

        sheet['B4'] = "Check CaptureRectangle for No Activate"
        if Tier.TryCaptureRectangle(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C4'] = "CaptureRectangle for No Activate is pass!!"
        else:
            sheet['C4'] = "CaptureRectangle for No Activate is fail!!"

        sheet['B5'] = "Check Record for No Activate"
        if Tier.TryRecord(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C5'] = "Record for No Activate is pass!!"
        else:
            sheet['C5'] = "Record for No Activate is fail!!"

        sheet['B6'] = "Check Stream for No Activate"
        if Tier.TryStream(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C6'] = "Stream for No Activate is pass!!"
        else:
            sheet['C6'] = "Stream for No Activate is fail!!"

        sheet['B7'] = "Check Save for No Activate"
        if Tier.TrySave(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C7'] = "Save for No Activate is pass!!"
        else:
            sheet['C7'] = "Save for No Activate is fail!!"

        sheet['B8'] = "Check SaveAs for No Activate"
        if Tier.TrySaveAs(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C8'] = "SaveAs for No Activate is pass!!"
        else:
            sheet['C8'] = "SaveAs for No Activate is fail!!"

        sheet['B9'] = "Check ExportBtn for No Activate"
        if Tier.TryExportBtn(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C9'] = "ExportBtn for No Activate is pass!!"
        else:
            sheet['C9'] = "ExportBtn for No Activate is fail!!"

        sheet['B10'] = "Check Mail for No Activate"
        if Tier.TryMail(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C10'] = "Mail for No Activate is pass!!"
        else:
            sheet['C10'] = "Mail for No Activate is fail!!"

        sheet['B11'] = "Check QRcodeShare for No Activate"
        if Tier.TryQRcodeShare(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C11'] = "QRcodeShare for No Activate is pass!!"
        else:
            sheet['C11'] = "QRcodeShare for No Activate is fail!!"

        sheet['B12'] = "Check StickyNote for No Activate"
        if Tier.TryStickyNote(self, "orange")[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C12'] = "StickyNote for No Activate is pass!!"
        else:
            sheet['C12'] = "StickyNote for No Activate is fail!!"

        sheet['B13'] = "Check Mathtool for No Activate"
        if Tier.TryMathtool(self, "compass")[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C13'] = "Mathtool for No Activate is pass!!"
        else:
            sheet['C13'] = "Mathtool for No Activate is fail!!"

        sheet['B14'] = "Check Camera for No Activate"
        if Tier.TryCamera(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C14'] = "Camera for No Activate is pass!!"
        else:
            sheet['C14'] = "Camera for No Activate is fail!!"

        sheet['B15'] = "Check YouTubeSearch for No Activate"
        if Tier.TryYouTubeSearch(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C15'] = "YouTubeSearch for No Activate is pass!!"
        else:
            sheet['C15'] = "YouTubeSearch for No Activate is fail!!"

        sheet['B16'] = "Check ImageSearch for No Activate"
        if Tier.TryImageSearch(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C16'] = "ImageSearch for No Activate is pass!!"
        else:
            sheet['C16'] = "ImageSearch for No Activate is fail!!"

        sheet['B17'] = "Check Throw for No Activate"
        if Tier.TryThrow(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C17'] = "Throw for No Activate is pass!!"
        else:
            sheet['C17'] = "Throw for No Activate is fail!!"

        sheet['B18'] = "Check PopQuiz for No Activate"
        if Tier.TryPopQuiz(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C18'] = "PopQuiz for No Activate is pass!!"
        else:
            sheet['C18'] = "PopQuiz for No Activate is fail!!"

        sheet['B19'] = "Check AIpen for No Activate"
        if Tier.TryAIpen(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C19'] = "AIpen for No Activate is pass!!"
        else:
            sheet['C19'] = "AIpen for No Activate is fail!!"

        sheet['B20'] = "Check FollowMeBackground for No Activate"
        if Tier.TryFollowMeBackground(self)[1] == "Whiteboard activation required" or "啟用提示":
            sheet['C20'] = "FollowMeBackground for No Activate is pass!!"
        else:
            sheet['C20'] = "FollowMeBackground for No Activate is fail!!"
        sheet['B21'] = "No Activate Finish"
    except (SystemExit, KeyboardInterrupt):
        raise
    except Exception:
        sheet['C2'] = "系統運行錯誤，請看log日誌"
        Tier_log.logger.error("系統運行錯誤，請重試", exc_info=True)
        Tier_log.logger.error("Finish")
    Tier_wd.save('./Excels/Tier.xlsx')


def TierPreloadCase(self):
    #Tier_wd = load_workbook('./Excels/Tier.xlsx')
    sheet = Tier_wd.worksheets[1]
    try:
        sheet['B2'] = "Check WirelessPresentation for No Preload"
        if Tier.TryWirelessPresentation(self)[0] == "Please sign in to use the feature.":
            sheet['C2'] = "WirelessPresentation for Tier1 is pass!!"
        else:
            sheet['C2'] = "WirelessPresentation for Tier1 is fail!!"

        sheet['B3'] = "Check CaptureAll for No Preload"
        if Tier.TryCaptureAll(self)[0] == "未出現Toast提示":
            sheet['C3'] = "CaptureAll for Tier1 is pass!!"
        else:
            sheet['C3'] = "CaptureAll for Tier1 is fail!!"

        sheet['B4'] = "Check CaptureRectangle for No Preload"
        if Tier.TryCaptureRectangle(self)[0] == "未出現Toast提示":
            sheet['C4'] = "CaptureRectangle for Tier1 is pass!!"
        else:
            sheet['C4'] = "CaptureRectangle for Tier1 is fail!!"

        sheet['B5'] = "Check Record for No Preload"
        if Tier.TryRecord(self)[0] == "未出現Toast提示":
            sheet['C5'] = "Record for Tier1 is pass!!"
        else:
            sheet['C5'] = "Record for Tier1 is fail!!"

        sheet['B5'] = "Check Stream for No Preload"
        if Tier.TryStream(self)[0] == "未出現Toast提示":
            sheet['C5'] = "Stream for Tier1 is pass!!"
        else:
            sheet['C5'] = "Stream for Tier1 is fail!!"

        sheet['B6'] = "Check Save for No Preload"
        if Tier.TrySave(self)[0] == "未出現Toast提示":
            sheet['C6'] = "Save for Tier1 is pass!!"
        else:
            sheet['C6'] = "Save for Tier1 is fail!!"

        sheet['B7'] = "Check SaveAs for No Preload"
        if Tier.TrySaveAs(self)[0] == "未出現Toast提示":
            sheet['C7'] = "SaveAs for Tier1 is pass!!"
        else:
            sheet['C7'] = "SaveAs for Tier1 is fail!!"

        sheet['B8'] = "Check ExportBtn for No Preload"
        if Tier.TryExportBtn(self)[0] == "未出現Toast提示":
            sheet['C8'] = "ExportBtn for Tier1 is pass!!"
        else:
            sheet['C8'] = "ExportBtn for Tier1 is fail!!"

        sheet['B9'] = "Check ExportPDF for No Preload"
        if Tier.TryExportPDF(self)[0] == "Please sign in to use the feature.":
            sheet['C9'] = "ExportPDF for Tier1 is pass!!"
        else:
            sheet['C9'] = "ExportPDF for Tier1 is fail!!"

        sheet['B10'] = "Check Mail for No Preload"
        if Tier.TryMail(self)[0] == "Please sign in to use the feature.":
            sheet['C10'] = "Mail for Tier1 is pass!!"
        else:
            sheet['C10'] = "Mail for Tier1 is fail!!"

        sheet['B11'] = "Check QRcodeShare for No Preload"
        if Tier.TryQRcodeShare(self)[0] == "Please sign in to use the feature.":
            sheet['C11'] = "QRcodeShare for Tier1 is pass!!"
        else:
            sheet['C11'] = "QRcodeShare for Tier1 is fail!!"

        sheet['B12'] = "Check StickyNote for No Preload"
        if Tier.TryStickyNote(self, "orange")[0] == "未出現Toast提示":
            sheet['C12'] = "StickyNote for Tier1 is pass!!"
        else:
            sheet['C12'] = "StickyNote for Tier1 is fail!!"

        sheet['B13'] = "Check Mathtool for No Preload"
        if Tier.TryMathtool(self, "compass")[0] == "未出現Toast提示":
            sheet['C13'] = "Mathtool for Tier1 is pass!!"
        else:
            sheet['C13'] = "Mathtool for Tier1 is fail!!"

        sheet['B14'] = "Check Camera for No Preload"
        if Tier.TryCamera(self)[0] == "未出現Toast提示":
            sheet['C14'] = "Camera for Tier1 is pass!!"
        else:
            sheet['C14'] = "Camera for Tier1 is fail!!"

        sheet['B15'] = "Check YouTubeSearch for No Preload"
        if Tier.TryYouTubeSearch(self)[0] == "Please sign in to use the feature.":
            sheet['C15'] = "YouTubeSearch for Tier1 is pass!!"
        else:
            sheet['C15'] = "YouTubeSearch for Tier1 is fail!!"

        sheet['B16'] = "Check ImageSearch for No Preload"
        if Tier.TryImageSearch(self)[0] == "Please sign in to use the feature.":
            sheet['C16'] = "ImageSearch for Tier1 is pass!!"
        else:
            sheet['C16'] = "ImageSearch for Tier1 is fail!!"

        sheet['B17'] = "Check Throw for No Preload"
        if Tier.TryThrow(self)[0] == "Please sign in to use the feature.":
            sheet['C17'] = "Throw for Tier1 is pass!!"
        else:
            sheet['C17'] = "Throw for Tier1 is fail!!"

        sheet['B18'] = "Check PopQuiz for No Preload"
        if Tier.TryPopQuiz(self)[0] == "Please sign in to use the feature.":
            sheet['C18'] = "PopQuiz for Tier1 is pass!!"
        else:
            sheet['C18'] = "PopQuiz for Tier1 is fail!!"

        sheet['B19'] = "Check AIpen for No Preload"
        if Tier.TryAIpen(self)[0] == "Please sign in to use the feature.":
            sheet['C19'] = "AIpen for Tier1 is pass!!"
        else:
            sheet['C19'] = "AIpen for Tier1 is fail!!"

        sheet['B20'] = "Check FollowMeBackground for No Preload"
        if Tier.TryFollowMeBackground(self)[0] == "Please sign in to use the feature.":
            sheet['C20'] = "FollowMeBackground for Tier1 is pass!!"
        else:
            sheet['C20'] = "FollowMeBackground for Tier1 is fail!!"
    except (SystemExit, KeyboardInterrupt):
        raise
    except Exception:
        sheet['C2'] = "系統運行錯誤，請看log日誌"
        Tier_log.logger.error("系統運行錯誤，請重試", exc_info=True)
        Tier_log.logger.error("Finish")
    Tier_wd.save('./Excels/Tier.xlsx')


def TierStandardCase(self):
    #Tier_wd = load_workbook('./Excels/Tier.xlsx')
    sheet = Tier_wd.worksheets[2]
    try:
        LoginAndActive.NormalLogin(
            self, "mvbaautotest1@gmail.com", "$Password1")
        time.sleep(10)

        sheet['B2'] = "Check WirelessPresentation for Standard"
        if Tier.TryWirelessPresentation(self)[1] == "Subscription upgrade required" or "訂閱資料":
            sheet['C2'] = "WirelessPresentation for Tier2 is pass!!"
        else:
            sheet['C2'] = "WirelessPresentation for Tier2 is fail!!"

        sheet['B3'] = "Check CaptureAll for Standard"
        if Tier.TryCaptureAll(self)[1] == "未出現提示":
            sheet['C3'] = "CaptureAll for Tier2 is pass!!"
        else:
            sheet['C3'] = "CaptureAll for Tier2 is fail!!"

        sheet['B4'] = "Check CaptureRectangle for Standard"
        if Tier.TryCaptureRectangle(self)[1] == "未出現提示":
            sheet['C4'] = "CaptureRectangle for Tier2 is pass!!"
        else:
            sheet['C4'] = "CaptureRectangle for Tier2 is fail!!"

        sheet['B5'] = "Check Record for Standard"
        if Tier.TryRecord(self)[1] == "未出現提示":
            sheet['C5'] = "Record for Tier2 is pass!!"
        else:
            sheet['C5'] = "Record for Tier2 is fail!!"

        sheet['B6'] = "Check Stream for Standard"
        if Tier.TryStream(self)[1] == "未出現提示":
            sheet['C6'] = "Stream for Tier2 is pass!!"
        else:
            sheet['C6'] = "Stream for Tier2 is fail!!"

        sheet['B7'] = "Check Save for Standard"
        if Tier.TrySave(self)[1] == "未出現提示":
            sheet['C7'] = "Save for Tier2 is pass!!"
        else:
            sheet['C7'] = "Save for Tier2 is fail!!"

        sheet['B8'] = "Check SaveAs for Standard"
        if Tier.TrySaveAs(self)[1] == "未出現提示":
            sheet['C8'] = "SaveAs for Tier2 is pass!!"
        else:
            sheet['C8'] = "SaveAs for Tier2 is fail!!"

        sheet['B9'] = "Check ExportBtn for Standard"
        if Tier.TryExportBtn(self)[1] == "未出現提示":
            sheet['C9'] = "ExportBtn for Tier2 is pass!!"
        else:
            sheet['C9'] = "ExportBtn for Tier2 is fail!!"

        sheet['B10'] = "Check ExportPDF for Standard"
        if Tier.TryExportPDF(self)[1] == "Subscription upgrade required" or "訂閱資料":
            sheet['C8'] = "ExportPDF for Tier2 is pass!!"
        else:
            sheet['C8'] = "ExportPDF for Tier2 is fail!!"

        sheet['C8'] = "Check Mail for Standard"
        if Tier.TryMail(self)[1] == "Subscription upgrade required" or "訂閱資料":
            sheet['C10'] = "Mail for Tier2 is pass!!"
        else:
            sheet['C10'] = "Mail for Tier2 is fail!!"

        sheet['B11'] = "Check QRcodeShare for Standard"
        if Tier.TryQRcodeShare(self)[1] == "Subscription upgrade required" or "訂閱資料":
            sheet['C11'] = "QRcodeShare for Tier2 is pass!!"
        else:
            sheet['C11'] = "QRcodeShare for Tier2 is fail!!"

        sheet['B12'] = "Check StickyNote for Standard"
        if Tier.TryStickyNote(self, "orange")[1] == "未出現提示":
            sheet['C12'] = "StickyNote for Tier2 is pass!!"
        else:
            sheet['C12'] = "StickyNote for Tier2 is fail!!"

        sheet['B13'] = "Check Mathtool for Standard"
        if Tier.TryMathtool(self, "compass")[1] == "未出現提示":
            sheet['C13'] = "Mathtool for Tier2 is pass!!"
        else:
            sheet['C13'] = "Mathtool for Tier2 is fail!!"

        sheet['B14'] = "Check Camera for Standard"
        if Tier.TryCamera(self)[1] == "未出現提示":
            sheet['C14'] = "Camera for Tier2 is pass!!"
        else:
            sheet['C14'] = "Camera for Tier2 is fail!!"

        sheet['B15'] = "Check YouTubeSearch for Standard"
        if Tier.TryYouTubeSearch(self)[1] == "未出現提示":
            sheet['C15'] = "YouTubeSearch for Tier2 is pass!!"
        else:
            sheet['C15'] = "YouTubeSearch for Tier2 is fail!!"

        sheet['B16'] = "Check ImageSearch for Standard"
        if Tier.TryImageSearch(self)[1] == "未出現提示":
            sheet['C16'] = "ImageSearch for Tier2 is pass!!"
        else:
            sheet['C16'] = "ImageSearch for Tier2 is fail!!"

        sheet['B17'] = "Check Throw for Standard"
        if Tier.TryThrow(self)[1] == "Subscription upgrade required" or "訂閱資料":
            sheet['C17'] = "Throw for Tier2 is pass!!"
        else:
            sheet['C17'] = "Throw for Tier2 is fail!!"

        sheet['B18'] = "Check PopQuiz for Standard"
        if Tier.TryPopQuiz(self)[1] == "Subscription upgrade required" or "訂閱資料":
            sheet['C18'] = "PopQuiz for Tier2 is pass!!"
        else:
            sheet['C18'] = "PopQuiz for Tier2 is fail!!"

        sheet['B19'] = "Check AIpen for Standard"
        if Tier.TryAIpen(self)[1] == "Subscription upgrade required" or "訂閱資料":
            sheet['C19'] = "AIpen for Tier2 is pass!!"
        else:
            sheet['C19'] = "AIpen for Tier2 is fail!!"

        sheet['B20'] = "Check FollowMeBackground for Standard"
        if Tier.TryFollowMeBackground(self)[1] == "未出現提示":
            sheet['C20'] = "FollowMeBackground for Tier2 is pass!!"
        else:
            sheet['C20'] = "FollowMeBackground for Tier2 is fail!!"
    except (SystemExit, KeyboardInterrupt):
        raise
    except Exception:
        sheet['C2'] = "系統運行錯誤，請看log日誌"
        Tier_log.logger.error("系統運行錯誤，請重試", exc_info=True)
        Tier_log.logger.error("Finish")
    Tier_wd.save('./Excels/Tier.xlsx')


def TierPremiumCase(self):
    #Tier_wd = load_workbook('./Excels/Tier.xlsx')
    sheet = Tier_wd.worksheets[3]
    try:
        LoginAndActive.NormalLogin(
            self, "mvbaautotest2@gmail.com", "$Password1")
        time.sleep(10)

        sheet['B2'] = "Check WirelessPresentation for Premium"
        if Tier.TryWirelessPresentation(self)[1] == "未出現提示":
            sheet['C2'] = "WirelessPresentation for Tier3 is pass!!"
        else:
            sheet['C2'] = "WirelessPresentation for Tier3 is fail!!"

        sheet['B3'] = "Check CaptureAll for Premium"
        if Tier.TryCaptureAll(self)[1] == "未出現提示":
            sheet['C3'] = "CaptureAll for Tier3 is pass!!"
        else:
            sheet['C3'] = "CaptureAll for Tier3 is fail!!"

        sheet['B4'] = "Check CaptureRectangle for Premium"
        if Tier.TryCaptureRectangle(self)[1] == "未出現提示":
            sheet['C4'] = "CaptureRectangle for Tier3 is pass!!"
        else:
            sheet['C4'] = "CaptureRectangle for Tier3 is fail!!"

        sheet['B5'] = "Check Record for Premium"
        if Tier.TryRecord(self)[1] == "未出現提示":
            sheet['C5'] = "Record for Tier3 is pass!!"
        else:
            sheet['C5'] = "Record for Tier3 is fail!!"

        sheet['B6'] = "Check Stream for Premium"
        if Tier.TryStream(self)[1] == "未出現提示":
            sheet['C6'] = "Stream for Tier3 is pass!!"
        else:
            sheet['C6'] = "Stream for Tier3 is fail!!"

        sheet['B7'] = "Check Save for Premium"
        if Tier.TrySave(self)[1] == "未出現提示":
            sheet['C7'] = "Save for Tier3 is pass!!"
        else:
            sheet['C7'] = "Save for Tier3 is fail!!"

        sheet['B8'] = "Check SaveAs for Premium"
        if Tier.TrySaveAs(self)[1] == "未出現提示":
            sheet['C8'] = "SaveAs for Tier3 is pass!!"
        else:
            sheet['C8'] = "SaveAs for Tier3 is fail!!"

        sheet['B9'] = "Check ExportBtn for Premium"
        if Tier.TryExportBtn(self)[1] == "未出現提示":
            sheet['C9'] = "ExportBtn for Tier3 is pass!!"
        else:
            sheet['C9'] = "ExportBtn for Tier3 is fail!!"

        sheet['B10'] = "Check ExportPDF for Premium"
        if Tier.TryExportPDF(self)[1] == "未出現提示":
            sheet['C10'] = "ExportPDF for Tier3 is pass!!"
        else:
            sheet['C10'] = "ExportPDF for Tier3 is fail!!"

        sheet['B11'] = "Check Mail for Premium"
        if Tier.TryMail(self)[1] == "未出現提示":
            sheet['C11'] = "Mail for Tier3 is pass!!"
        else:
            sheet['C11'] = "Mail for Tier3 is fail!!"

        sheet['B12'] = "Check QRcodeShare for Premium"
        if Tier.TryQRcodeShare(self)[1] == "Share the Whiteboard session with the QR Code" or "掃描 QR code 以分享畫布":
            sheet['C12'] = "QRcodeShare for Tier3 is pass!!"
        else:
            sheet['C12'] = "QRcodeShare for Tier3 is fail!!"

        sheet['B13'] = "Check StickyNote for Premium"
        if Tier.TryStickyNote(self, "orange")[1] == "未出現提示":
            sheet['C13'] = "StickyNote for Tier3 is pass!!"
        else:
            sheet['C13'] = "StickyNote for Tier3 is fail!!"

        sheet['B14'] = "Check Mathtool for Premium"
        if Tier.TryMathtool(self, "compass")[1] == "未出現提示":
            sheet['C14'] = "Mathtool for Tier3 is pass!!"
        else:
            sheet['C14'] = "Mathtool for Tier3 is fail!!"

        sheet['B15'] = "Check Camera for Premium"
        if Tier.TryCamera(self)[1] == "未出現提示":
            sheet['C15'] = "Camera for Tier3 is pass!!"
        else:
            sheet['C15'] = "Camera for Tier3 is fail!!"

        sheet['B16'] = "Check YouTubeSearch for Premium"
        if Tier.TryYouTubeSearch(self)[1] == "未出現提示":
            sheet['C16'] = "YouTubeSearch for Tier3 is pass!!"
        else:
            sheet['C16'] = "YouTubeSearch for Tier3 is fail!!"

        sheet['B17'] = "Check ImageSearch for Premium"
        if Tier.TryImageSearch(self)[1] == "未出現提示":
            sheet['C17'] = "ImageSearch for Tier3 is pass!!"
        else:
            sheet['C17'] = "ImageSearch for Tier3 is fail!!"

        sheet['B18'] = "Check Throw for Premium"
        if Tier.TryThrow(self)[1] == "未出現提示":
            sheet['C18'] = "Throw for Tier3 is pass!!"
        else:
            sheet['C18'] = "Throw for Tier3 is fail!!"

        sheet['B19'] = "Check PopQuiz for Premium"
        if Tier.TryPopQuiz(self)[1] == "未出現提示":
            sheet['C19'] = "PopQuiz for Tier3 is pass!!"
        else:
            sheet['C19'] = "PopQuiz for Tier3 is fail!!"

        sheet['B20'] = "Check AIpen for Premium"
        if Tier.TryAIpen(self)[1] == "未出現提示":
            sheet['C20'] = "AIpen for Tier3 is pass!!"
        else:
            sheet['C20'] = "AIpen for Tier3 is fail!!"

        sheet['B21'] = "Check FollowMeBackground for Premium"
        if Tier.TryFollowMeBackground(self)[1] == "未出現提示":
            sheet['C21'] = "FollowMeBackground for Tier3 is pass!!"
        else:
            sheet['C21'] = "FollowMeBackground for Tier3 is fail!!"
    except (SystemExit, KeyboardInterrupt):
        raise
    except Exception:
        sheet['C2'] = "系統運行錯誤，請看log日誌"
        Tier_log.logger.error("系統運行錯誤，請重試", exc_info=True)
        Tier_log.logger.error("Finish")
    Tier_wd.save('./Excels/Tier.xlsx')


def TierEntityCase(self):
    #Tier_wd = load_workbook('./Excels/Tier.xlsx')
    sheet = Tier_wd.worksheets[4]
    try:
        LoginAndActive.NormalLogin(
            self, "mvbaautotest4@gmail.com", "$Password1")
        time.sleep(10)
        sheet['B2'] = "Check WirelessPresentation for Entity"
        if Tier.TryWirelessPresentation(self)[1] == "未出現提示":
            sheet['C2'] = "WirelessPresentation for Tier4 is pass!!"
        else:
            sheet['C2'] = "WirelessPresentation for Tier4 is fail!!"

        sheet['B3'] = "Check CaptureAll for Entity"
        if Tier.TryCaptureAll(self)[1] == "未出現提示":
            sheet['C3'] = "CaptureAll for Tier4 is pass!!"
        else:
            sheet['C3'] = "CaptureAll for Tier4 is fail!!"

        sheet['B3'] = "Check FollowMeBackground for Entity"
        if Tier.TryCaptureRectangle(self)[1] == "未出現提示":
            sheet['C3'] = "CaptureRectangle for Tier4 is pass!!"
        else:
            sheet['C3'] = "CaptureRectangle for Tier4 is fail!!"

        sheet['B4'] = "Check Record for Entity"
        if Tier.TryRecord(self)[1] == "未出現提示":
            sheet['C4'] = "Record for Tier4 is pass!!"
        else:
            sheet['C4'] = "Record for Tier4 is fail!!"

        sheet['B5'] = "Check Stream for Entity"
        if Tier.TryStream(self)[1] == "未出現提示":
            sheet['C5'] = "Stream for Tier4 is pass!!"
        else:
            sheet['C5'] = "Stream for Tier4 is fail!!"

        sheet['B6'] = "Check Save for Entity"
        if Tier.TrySave(self)[1] == "未出現提示":
            sheet['C6'] = "Save for Tier4 is pass!!"
        else:
            sheet['C6'] = "Save for Tier4 is fail!!"

        sheet['B7'] = "Check SaveAs for Entity"
        if Tier.TrySaveAs(self)[1] == "未出現提示":
            sheet['C7'] = "SaveAs for Tier4 is pass!!"
        else:
            sheet['C7'] = "SaveAs for Tier4 is fail!!"

        sheet['B8'] = "Check ExportBtn for Entity"
        if Tier.TryExportBtn(self)[1] == "未出現提示":
            sheet['C8'] = "ExportBtn for Tier4 is pass!!"
        else:
            sheet['C8'] = "ExportBtn for Tier4 is fail!!"

        sheet['B9'] = "Check ExportPDF for Entity"
        if Tier.TryExportPDF(self)[1] == "未出現提示":
            sheet['C9'] = "ExportPDF for Tier4 is pass!!"
        else:
            sheet['C9'] = "ExportPDF for Tier4 is fail!!"

        sheet['B10'] = "Check Mail for Entity"
        if Tier.TryMail(self)[1] == "未出現提示":
            sheet['C10'] = "Mail for Tier4 is pass!!"
        else:
            sheet['C10'] = "Mail for Tier4 is fail!!"

        sheet['B11'] = "Check QRcodeShare for Entity"
        if Tier.TryQRcodeShare(self)[1] == "Share the Whiteboard session with the QR Code" or "掃描 QR code 以分享畫布":
            sheet['C11'] = "QRcodeShare for Tier4 is pass!!"
        else:
            sheet['C11'] = "QRcodeShare for Tier4 is fail!!"

        sheet['B12'] = "Check StickyNote for Entity"
        if Tier.TryStickyNote(self, "orange")[1] == "未出現提示":
            sheet['C12'] = "StickyNote for Tier4 is pass!!"
        else:
            sheet['C12'] = "StickyNote for Tier4 is fail!!"

        sheet['B13'] = "Check Mathtool for Entity"
        if Tier.TryMathtool(self, "compass")[1] == "未出現提示":
            sheet['C13'] = "Mathtool for Tier4 is pass!!"
        else:
            sheet['C13'] = "Mathtool for Tier4 is fail!!"

        sheet['B14'] = "Check Camera for Entity"
        if Tier.TryCamera(self)[1] == "未出現提示":
            sheet['C14'] = "Camera for Tier4 is pass!!"
        else:
            sheet['C14'] = "Camera for Tier4 is fail!!"

        sheet['B15'] = "Check YouTubeSearch for Entity"
        if Tier.TryYouTubeSearch(self)[1] == "未出現提示":
            sheet['C15'] = "YouTubeSearch for Tier4 is pass!!"
        else:
            sheet['C15'] = "YouTubeSearch for Tier4 is fail!!"

        sheet['B16'] = "Check ImageSearch for Entity"
        if Tier.TryImageSearch(self)[1] == "未出現提示":
            sheet['C16'] = "ImageSearch for Tier4 is pass!!"
        else:
            sheet['C16'] = "ImageSearch for Tier4 is fail!!"

        sheet['B17'] = "Check Throw for Entity"
        if Tier.TryThrow(self)[1] == "未出現提示":
            sheet['C17'] = "Throw for Tier4 is pass!!"
        else:
            sheet['C17'] = "Throw for Tier4 is fail!!"

        sheet['B18'] = "Check PopQuiz for Entity"
        if Tier.TryPopQuiz(self)[1] == "未出現提示":
            sheet['C18'] = "PopQuiz for Tier4 is pass!!"
        else:
            sheet['C18'] = "PopQuiz for Tier4 is fail!!"

        sheet['B19'] = "Check AIpen for Entity"
        if Tier.TryAIpen(self)[1] == "未出現提示":
            sheet['C19'] = "AIpen for Tier4 is pass!!"
        else:
            sheet['C19'] = "AIpen for Tier4 is fail!!"

        sheet['B20'] = "Check FollowMeBackground for Entity"
        if Tier.TryFollowMeBackground(self)[1] == "未出現提示":
            sheet['C20'] = "FollowMeBackground for Tier4 is pass!!"
        else:
            sheet['C20'] = "FollowMeBackground for Tier4 is fail!!"

        sheet['B21'] = "Check Clips for Entity"
        if Tier.TryClips(self):
            sheet['C21'] = "Clips for Tier4 is pass!!"
        else:
            sheet['C21'] = "Clips for Tier4 is fail!!"
    except (SystemExit, KeyboardInterrupt):
        raise
    except Exception:
        sheet['C2'] = "系統運行錯誤，請看log日誌"
        Tier_log.logger.error("系統運行錯誤，請重試", exc_info=True)
        Tier_log.logger.error("Finish")
    Tier_wd.save('./Excels/Tier.xlsx')
