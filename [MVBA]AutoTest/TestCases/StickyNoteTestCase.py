from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import time
from Library.LibExcel import LibExcel
from UnitFunction import Canvas, StickyNote, MagicBox, Screenshot, Compare, log, LoginAndActive
import sys
import os
sys.path.append(os.getcwd())

getTime = datetime.datetime.now().strftime(' %Y%m%d %H_%M_%S')
StickyNote_w = Workbook()
w_s = StickyNote_w.active
w_s.title = "Pen"
w_s1 = StickyNote_w.create_sheet(index=1, title="Text")
for w in range(2):
    sheet = StickyNote_w.worksheets[w]
    sheet['A1'] = "時間"
    sheet['B1'] = "步驟"
    sheet['C1'] = "結果"
    sheet['D1'] = '圖片'
    sheet.column_dimensions['A'].width = 18.0  # 調整列寬
    sheet.column_dimensions['B'].width = 50.0
    sheet.column_dimensions['C'].width = 50.0
    #sheet.column_dimensions['D'].width = 48
StickyNote_w.save('./Output/Excels/StickyNote' + getTime + '.xlsx')
StickyNote_log = log.Logger(
    './Output/Logs/StickyNote'+datetime.datetime.now().strftime(' %Y%m%d %H_%M_%S') + '.log', level='info')


def TestStickyNote_Pen(self):
    try:
        allStep = LibExcel.TestCase_steps(
            './TestCases/case document/StickyNote', 0)
        #sheet = StickyNote_wd.worksheets[0]
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, step=allStep[0])
        #sheet['A2'] = "Case1-1: 實作StickyNote的Pen模式"
        #sheet['B2'] = "Open MagicBox"
        MagicBox.OpenMagicBox(self)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, step=allStep[1])
        #sheet['B3'] = "Click Stickynote icon"
        StickyNote.ClickStickynote(self)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, step=allStep[2])
        #sheet['B4'] = "Click Stickynote Color icon"
        StickyNote.StickynoteColor(self, 2)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, step=allStep[3])
        #sheet['B5'] = "Click Pen icon"
        StickyNote.SelectPen(self)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, step=allStep[4])
        #sheet['B6'] = "Click Color icon"
        StickyNote.PenColor(self, "Red")
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, step=allStep[5])
        #sheet['B7'] = "Ajust Pen Thickness"
        StickyNote.PenThickness(self, 8)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, step=allStep[6])
        #sheet['B8'] = "Draw a 6"
        my_draw = [[1000, 350], [850, 350], [850, 650],
                   [1000, 650], [1000, 500], [850, 500]]
        Canvas.Swipe(self, my_draw)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, step=allStep[7])
        #sheet['B9'] = "Screenshot StickyNote Canvas"
        Screenshot.screenshotStickyNoteCanvas(
            self, "./TestCases/ScreenShots/StickyNote/Canvas_Pen.png")
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, step=allStep[8])
        #sheet['B10'] = "Screenshot StickyNote Bar"
        Screenshot.screenshotStickyNoteBar(
            self, "./TestCases/ScreenShots/StickyNote/Bar_Pen.png")
        time.sleep(3)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, step=allStep[9])
        #sheet['B11'] = "Close StickyNote"
        StickyNote.CloseStickynote(self)
        #sheet['B12'] = "比對Screenshot StickyNote Canvas"
        if not Compare.isImgEqual("./TestCases/ScreenShots/StickyNote/Canvas_Pen.png", "./TestCases/Samples/StickyNote/Canvas_Pen.png"):
            LibExcel.export_table(
                './Excels/StickyNote' + getTime, step=allStep[10], result="Failed")
            #sheet['C12'] = "Pen Case: Canvas Fail!"
        else:
            #sheet['C12'] = "Pen Case: Canvas Pass!"
            LibExcel.export_table(
                './Excels/StickyNote' + getTime, step=allStep[10], result="Pass")
        #sheet['B13'] = "比對Screenshot StickyNote Bar"
        if not Compare.isImgEqual("./TestCases/ScreenShots/StickyNote/Bar_Pen.png", "./TestCases/Samples/StickyNote/Bar_Pen.png"):
            #sheet['C13'] = "Fail!"
            LibExcel.export_table(
                './Excels/StickyNote' + getTime, step=allStep[11], result="Failed")
        else:
            #sheet['C13'] = "Pass!"
            LibExcel.export_table(
                './Excels/StickyNote' + getTime, step=allStep[11], result="Pass")
    except (SystemExit, KeyboardInterrupt):
        raise
    except Exception:
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, step=allStep[0], result="系統運行錯誤，請看log日誌")
        #sheet['C2'] = "系統運行錯誤，請看log日誌"
        StickyNote_log.logger.error("系統運行錯誤", exc_info=True)
        StickyNote_log.logger.info("Finish")
    #StickyNote_wd.save('./Excels/StickyNote' + getTime + '.xlsx')


def TestStickyNote_Text(self):
    try:
        allStep = LibExcel.TestCase_steps(
            './TestCases/case document/StickyNote', 1)
        #sheet = StickyNote_wd.worksheets[1]
        #sheet['A2'] = "Case1-2: 實作StickyNote的Text模式"
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[0])
        #sheet['B2'] = "Open MagicBox"
        MagicBox.OpenMagicBox(self)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[1])
        #sheet['B3'] = "Click Stickynote icon"
        StickyNote.ClickStickynote(self)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[2])
        #sheet['B4'] = "Click Stickynote Color icon"
        StickyNote.StickynoteColor(self, 4)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[3])
        #sheet['B5'] = "Click Stickynote T icon"
        StickyNote.SelectText(self)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[4])
        #sheet['B6'] = "Change Font"
        StickyNote.ChangeFont(self, 6)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[5])
        #sheet['B7'] = "Change Size"
        StickyNote.ChangeSize(self, 20)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[6])
        #sheet['B8'] = "Click Text Bold icon"
        StickyNote.TextBold(self)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[7])
        #sheet['B9'] = "Click Text Italic icon"
        StickyNote.TextItalic(self)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[8])
        #sheet['B10'] = "Click Text UnderLine icon"
        StickyNote.TextUnderLine(self)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[9])
        #sheet['B11'] = "輸入文字: teststickynote"
        StickyNote.TypeText(self, "teststickynote")
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[10])
        #sheet['B12'] = "Screenshot StickyNote Bar"
        Screenshot.screenshotStickyNoteBar(
            self, "./TestCases/ScreenShots/StickyNote/Bar_Text.png")
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[11])
        #sheet['B13'] = "返回Pen模式防止游標出現"
        StickyNote.SelectPen(self)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[12])
        #sheet['B14'] = "Screenshot StickyNote Canvas"
        Screenshot.screenshotStickyNoteCanvas(
            self, "./TestCases/ScreenShots/StickyNote/Canvas_Text.png")
        time.sleep(3)
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[13])
        #sheet['B15'] = "Close StickyNote"
        StickyNote.CloseStickynote(self)
        #sheet['B16'] = "比對Screenshot StickyNote Canvas"
        if not Compare.isImgEqual("./TestCases/ScreenShots/StickyNote/Canvas_Text.png", "./TestCases/Samples/StickyNote/Canvas_Text.png"):
            LibExcel.export_table(
                './Excels/StickyNote' + getTime, 1, step=allStep[14], result="Failed")
            #sheet['C16'] = "Text Case: Canvas Fail!"
        else:
            LibExcel.export_table(
                './Excels/StickyNote' + getTime, 1, step=allStep[14], result="Pass")
            #sheet['C16'] = "Text Case: Canvas Pass!"
        #sheet['B17'] = "比對Screenshot StickyNote Bar"
        if not Compare.isImgEqual("./TestCases/ScreenShots/StickyNote/Bar_Text.png", "./TestCases/Samples/StickyNote/Bar_Text.png"):
            LibExcel.export_table(
                './Excels/StickyNote' + getTime, 1, step=allStep[15], result="Failed")
            #sheet['C17'] = "Text Case: Bar Fail!"
        else:
            LibExcel.export_table(
                './Excels/StickyNote' + getTime, 1, step=allStep[15], result="Pass")
            #sheet['C17'] = "Text Case: Bar Pass!"
        #sheet['B18'] = "Case1-2 Finish"
    except (SystemExit, KeyboardInterrupt):
        raise
    except Exception:
        LibExcel.export_table(
            './Excels/StickyNote' + getTime, 1, step=allStep[0], result="系統運行錯誤，請看log日誌")
        #sheet['C2'] = "系統運行錯誤，請看log日誌"
        StickyNote_log.logger.error("系統運行錯誤，請重試", exc_info=True)
        StickyNote_log.logger.info("Finish")
